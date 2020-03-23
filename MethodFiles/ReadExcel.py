import requests
import openpyxl
from openpyxl import Workbook

File_Path="D:\Programming\Application Security\coe-application-security\Data Files\Data.xlsx"
Excel_Location= "D:\Programming\Application Security\coe-application-security\DataFiles\API.xlsx"
Excel_Sheet_Name = 'Sheet1'
ModuleName = "Technicalskills_POST"

def read_specific_cell_values(Excel_Location, Excel_Sheet_Name):
    try:
        Excel_WorkBook = openpyxl.load_workbook(Excel_Location)
        for SheetName in Excel_WorkBook.worksheets:
            if SheetName.title == Excel_Sheet_Name:
                print("Sheet Found")
                ActiveWorkSheet = Excel_WorkBook[SheetName.title]
                print("Title of Sheet = " + ActiveWorkSheet.title)
                # Read Cell Value
                print(ActiveWorkSheet['A2'].value)
                print(ActiveWorkSheet['B2'].value)
                break
            else:
                print("Sheet Not Found")
    except:
        print("Failed for execute")


def read_located_cell_values(Excel_Location, Excel_Sheet_Name):
    Excel_WorkBook = openpyxl.load_workbook(Excel_Location)
    for SheetName in Excel_WorkBook.worksheets:
        if SheetName.title == Excel_Sheet_Name:
            print("Sheet Found")
            ActiveWorkSheet = Excel_WorkBook[SheetName.title]
            print("Title of Sheet = " + ActiveWorkSheet.title)
            Expected_Cell_Value = SheetName.cell(2, 1)
            print(Expected_Cell_Value.value)

def read_refernce_cell_values(Excel_Location, Excel_Sheet_Name):
    Excel_WorkBook = openpyxl.load_workbook(Excel_Location)
    for SheetName in Excel_WorkBook.worksheets:
        if SheetName.title == Excel_Sheet_Name:
            print("Sheet Found")
            ActiveWorkSheet = Excel_WorkBook[SheetName.title]
            print("Title of Sheet = " + ActiveWorkSheet.title)
            Expected_Cell_Value = SheetName.cell(column=1, row=2)
            print(Expected_Cell_Value.value)


def read_number_of_rows_and_columns(Excel_Location, Excel_Sheet_Name):
    Excel_WorkBook = openpyxl.load_workbook(Excel_Location)
    for SheetName in Excel_WorkBook.worksheets:
        if SheetName.title == Excel_Sheet_Name:
            print("Sheet Found")
            ActiveWorkSheet = Excel_WorkBook[SheetName.title]
            print("Title of Sheet = " + ActiveWorkSheet.title)
    rows = SheetName.max_row
    columns = SheetName.max_column
    print("Total Rows are = " + str(rows))
    print("Total Columns are = " + str(columns))


def read_first_row_values(Excel_Location, Excel_Sheet_Name):
    Excel_WorkBook = openpyxl.load_workbook(Excel_Location)
    for SheetName in Excel_WorkBook.worksheets:
        if SheetName.title == Excel_Sheet_Name:
            print("Worksheet Found = " + SheetName.title)
            ActiveWorkSheet = Excel_WorkBook[SheetName.title]
            print("Title of Sheet = " + ActiveWorkSheet.title)

            RowLength = SheetName.max_row
            ColumnLength = SheetName.max_column
            print("Number of Payloads for " + str(SheetName.title) + " = " + str(RowLength - 1))

            for i in range(2, RowLength + 1):
                RowContents = SheetName.cell(row=i, column=1)
                print("Row "+ str(RowContents.row-1) + " = "+ RowContents.value, end="" + "\n")
                return RowContents.value


