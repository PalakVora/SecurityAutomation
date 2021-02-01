import openpyxl,traceback,json


#========================== FIND API NAME FROM EXCEL SHEET ======================================

def find_module_name(sheetname,module_name):
    rowlength = sheetname.max_row
    #ColumnLength = Sheetname.max_column
    for i in range(1, rowlength + 2):
        rowcontents = sheetname.cell(row=i, column=1)
        print("Row " + str(rowcontents.row) + " = " + str(rowcontents.value))
        if ((rowcontents.value) == module_name):
            print("Module Found : " + str(rowcontents.value) + " in Row - " + str(rowcontents.row) + " of Worksheet - " + sheetname.title)
            return rowcontents
        else:
            continue
        try:
            print(str((sheetname["A" + str(rowcontents.row)]).value) + "\n That API name section")
        except:
            print("Error in getting API Name")
            traceback.print_stack()
        break
 # ------------------------------------ Read Base URL ------------------------             
def read_base_url(sheetname,rowcontents,excel_workBook):
        try:
            request_base_url = (sheetname["D" + str(rowcontents.row)])
            print("BaseURL = " + str(request_base_url.value) + "\n")
            return str(request_base_url.value)
           
        except:
            print("Error in reading Request Base URL from Excel File")
            excel_workBook.close() 

# ------------------------- Read BODY---------------------
def read_body(sheetname,rowcontents,excel_workbook):
    try:
        body_row = (sheetname["F" + str(rowcontents.row)])
        print("Body ROW",body_row.value)
        return str(body_row.value)
    except:
        print("Error in reading Request Body from Excel File")
        excel_workbook.close()

# ----------------------- Get Cookie---------------------------------
def read_cookie(sheetname,rowcontents,excel_workbook):
    try:
        cookie_row = (sheetname["H" + str(rowcontents.row)])
        print("Body ROW",cookie_row.value)
        return str(cookie_row.value)
    except:
        print("Error in reading Request Cookie from Excel File")
        excel_workbook.close()

# ------------------------------ Get Header -------------------------
def read_header(sheetname,rowcontents,excel_workbook):
    try:
        header_row = (sheetname["G" + str(rowcontents.row)])
        return str(header_row.value)
    except:
        print("Error in reading Request Header from Excel File")
        excel_workbook.close()      


# ------------------------------ Get Protocol -------------------------
def read_protocol(sheetname,rowcontents,excel_workbook):
        try:
            request_protocol = (sheetname["C" + str(rowcontents.row)])
            print(request_protocol.value)
            return str(request_protocol.value)
        except:
            print("Error in reading Request Protocol from Excel File")
            excel_workbook.close()     

# ------------------------------ Get Relative URL -------------------------
def read_relative_url(sheetname,rowcontents,excel_workbook):   
        try:
             request_relative_url = (sheetname["E" + str(rowcontents.row)])
             print("RelativeURL = " + str(request_relative_url.value) + "\n")
             return str(request_relative_url.value)
        except:
            print("Error in reading Request Relative URL from Excel File")
            excel_workbook.close()

# ------------------------------ Get Verb -------------------------
def read_method(sheetname,rowcontents,excel_workbook):
        try:
            http_method = (sheetname["B" + str(rowcontents.row)])
            print("ROw" + str(rowcontents.row))
            print(http_method.value)
            return str(http_method.value)
        except Exception as error:
            print("Error in reading HTTP Method from Excel File")
            print(error)
            excel_workbook.close()
            return 0
                        
